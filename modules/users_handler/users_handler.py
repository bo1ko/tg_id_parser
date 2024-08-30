from modules.actions import Actions
from modules.search import Search

import openpyxl
import os


class UsersHandler(Actions):
    
    def __init__(self, driver, wait):
        super().__init__(driver, wait)
    
    def _get_usernames(self):
        workbook = openpyxl.load_workbook('excel/username.xlsx')        
        sheet = workbook.active
        usernames_arr = []

        for row in sheet.iter_rows(values_only=True):
            if row[0] != None:
                if row[0][0] == '@':
                    usernames_arr.append(row[0][1:])
        
        return usernames_arr
    
    def _fill_excel_by_user_id(self, users_dict):
        if os.path.exists('excel/usernames with id.xlsx'):
            os.remove('excel/usernames with id.xlsx')

        workbook = openpyxl.Workbook()   
        sheet = workbook.active

        sheet.append(['username', 'user_id'])
        for key, value in users_dict.items():
            sheet.append([key, value])
        
        workbook.save('excel/usernames with id.xlsx')

    def find_users_id(self):
        usernames_arr = self._get_usernames()
        search = Search(self._driver, self._wait)
        users_dict = {}

        print('Починаю шукати ID юзерів\n')
        for username in usernames_arr:
            user_id = search._user_search(username)

            print(f'{username}: {user_id}')
            users_dict[username] = user_id
    
        self._fill_excel_by_user_id(users_dict)

        print('\nГотово!')
